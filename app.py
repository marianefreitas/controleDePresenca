from flask import Flask, render_template
from tkinter.filedialog import askopenfilename
import pandas as pd
import openpyxl
import sqlite3
import matplotlib.pyplot as plt
import mpld3

Database_Name = "postgres://gnwaijkgmxivpg:ac715d6425dd87b40b708d924d7b73b14381a36ab9db14acd6d192961dd5a86b@ec2-107-21-67-46.compute-1.amazonaws.com:5432/dal8ggfqd1j4eg"
Table_Name = "lista_de_presenca"

Query_CreateTable = f"""
    CREATE TABLE IF NOT EXISTS {Table_Name} (
    id INTEGER PRIMARY KEY AUTOINCREMENT,
    created TIMESTAMP NOT NULL DEFAULT CURRENT_TIMESTAMP,
    school TEXT,
    subject TEXT,
    class TEXT,
    student_name TEXT,
    student_ra TEXT,
    date_activity TEXT,
    date_reference TEXT,
    student_attend TEXT,
    student_status TEXT
);"""


def make_autopct(values):
    def my_autopct(pct):
        total = sum(values)
        val = int(round(pct*total/100.0))
        return '{p:.0f}%  ({v:d})'.format(p=pct,v=val)
    return my_autopct


app = Flask(__name__)

@app.route('/')
def index():
    return render_template('index.html')


@app.route('/upload')
def upload_file():
    file_name = askopenfilename()
    print(" filename => ", file_name)

    workbook = openpyxl.load_workbook(file_name)
    worksheet = workbook["Lista Presença_Alunos"]

    file_school = worksheet['E6'].value
    file_subject = worksheet['E8'].value
    file_class = worksheet['X8'].value

    date_reference = ""
    first_interaction = True
    for row in range(14, 70):

        for col in range(5,74):

            if worksheet.cell(row=row, column=2).value is not None and worksheet.cell(row=13, column=col).value is not None:

                if date_reference == "":
                    date_reference = str(worksheet.cell(row=13, column=col).value)[:7]

                if worksheet.cell(row=row, column=col).value == "P" or worksheet.cell(row=row, column=col).value == "FJ":
                    student_status = "P"
                else:
                    student_status = "F"

                if first_interaction == True:
                    row_data = {
                        "school": [file_school],
                        "subject": [file_subject],
                        "class": [file_class],
                        "student_name": [worksheet.cell(row=row, column=2).value],
                        "student_ra": [worksheet.cell(row=row, column=3).value],
                        "date_activity": [worksheet.cell(row=13, column=col).value],
                        "date_reference": [str(worksheet.cell(row=13, column=col).value)[:7]],
                        "student_attend": [worksheet.cell(row=row, column=col).value],
                        "student_status": [student_status]
                    }
                    file_data = pd.DataFrame(row_data)
                    first_interaction = False
                else:
                    row_data = {
                        "school": file_school,
                        "subject": file_subject,
                        "class": file_class,
                        "student_name": worksheet.cell(row=row, column=2).value,
                        "student_ra": worksheet.cell(row=row, column=3).value,
                        "date_activity": worksheet.cell(row=13, column=col).value,
                        "date_reference": str(worksheet.cell(row=13, column=col).value)[:7],
                        "student_attend": worksheet.cell(row=row, column=col).value,
                        "student_status": student_status
                    }
                    file_data.loc[len(file_data)] = row_data

    print(" file_data => \n", file_data.to_string(index=True, header=True))

    connection = sqlite3.connect(Database_Name)
    connection.execute(Query_CreateTable)
    connection.execute(f"DELETE FROM {Table_Name} WHERE class = '{file_class}' AND date_reference = '{date_reference}'")
    file_data.to_sql(Table_Name, connection, index=False, if_exists='append')
    connection.commit()
    print(" database => \n", pd.read_sql(f"SELECT * FROM {Table_Name}", connection).to_string(index=False, header=True))
    connection.close()

    return render_template('upload.html')


@app.route('/report')
def create_reports():

    connection = sqlite3.connect(Database_Name)

    query_report = f"""
        SELECT class, COUNT(DISTINCT student_name) AS student_count, 
            (CAST(CAST(COUNT(IIF(student_status = 'F', 1, NULL)) AS FLOAT)/CAST(COUNT(student_status) AS FLOAT)*100 AS INTEGER) || '%') AS absence_percent 
        FROM {Table_Name} 
        GROUP BY class    
    """
    report1 = pd.read_sql(query_report, connection)
    report1 = report1.rename({'class': 'Sala de Aula', 'student_count': 'Quantidade de Alunos','absence_percent': 'Percentual de Faltas'}, axis=1)

    query_report = f"""
            SELECT class, COUNT(student_status) AS absence_no, student_name
            FROM {Table_Name} 
            WHERE student_status = 'F' 
            GROUP BY class, student_name 
            ORDER BY COUNT(student_status) DESC, student_name 
            LIMIT 5
        """
    report2 = pd.read_sql(query_report, connection)
    report2 = report2.rename({'class': 'Sala de Aula', 'student_name': 'Nome do Aluno', 'absence_no': 'Número de Faltas'},axis=1)

    query_report = f"""
                SELECT class, COUNT(student_status) AS absence_no
                FROM {Table_Name}
                WHERE student_status = 'F'
                GROUP BY class
            """
    report3 = pd.read_sql(query_report, connection)
    figure1 = plt.figure(frameon=False, clear=True)
    figure1.subplots_adjust(bottom=0)
    figure1.subplots_adjust(top=1)
    figure1.subplots_adjust(right=1)
    figure1.subplots_adjust(left=0)
    plt.pie(report3["absence_no"], labels=report3["class"], autopct=make_autopct(report3["absence_no"]), textprops={'fontsize': 15})
    figure1 = mpld3.fig_to_html(figure1, no_extras=True)

    return render_template('report.html', table1=[report1.to_html(classes='data', index=False, header="true")], table2=[report2.to_html(classes='data', index=False, header="true")], figure1=[figure1])


if __name__ == "__main__":
    app.run(debug=True)